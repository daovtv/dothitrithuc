from openpyxl import load_workbook

class Excel:
    filename = ''
    def __init__(self,filename):
        self.filename = filename
    def open_workbook(self):
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active
    def save(self,filename):
        self.wb.save(filename)
    def select_sheet(self,name:str):
        self.ws = self.wb[name]
        return self.ws
    def select_sheet_index(self,index:int):
        sheets = self.wb.sheetnames
        self.ws= self.wb[sheets[index]]
        return self.ws
    def nrow(self,sheet, col1=9 ): # ten thuoc tinh        
        rows = 1        
        for sr in sheet.iter_rows():
            if type(sr[col1].value) !=type(None) and sr[col1].value!='':
              rows=rows+1
              
        return rows
    def ncol(self,sheet,row1=1):   # tieu de     
        cols = 1        
        for sr in sheet.iter_cols():
            if type(sr[row1].value) !=type(None) and sr[row1].value!='':
              cols=cols+1
        return cols
    def print_sheet (self,index):
        sheets = self.wb.sheetnames 
        sheet = self.wb[sheets[index]]
        print ('Sheet:',sheet.title)
        json = {'Schema':'Luocdo'}
        rows=1
        col1=9 # ten thuoc tinh
        for sr in sheet.iter_rows():
            if type(sr[col1]) !=type(None):
              rows=rows+1
        print(rows)
        for row in range(1,rows):          
          if row > 1:
             col=1
             cparent = 4
             if row == 2:                
                    json['Object']=sheet.cell(row,col).value.strip() # ten doi tuong
                    if (type(sheet.cell(row,cparent).value))!= type(None):
                        json['Parent']= sheet.cell(row,cparent).value.strip() # ten doi tuong cha
                    else:
                        json['Parent']=''
                    json['Property']={}
                    json['Rel']= {}
             else:
                col1=8+1 # ten thuoc tinh
                col2=6+1 # Property 
                col3=7+1 # Relation
                if type(sheet.cell(row,col2).value)!= type(None):
                    json['Property'][sheet.cell(row,col1).value]=sheet.cell(row,col2).value.strip() # thuoc tinh la kieu du lieu chuan
                    
                if type(sheet.cell(row,col3).value)!= type(None):
                    if sheet.cell(row,col3).value.strip()==json['Object'].strip():
                        json['Property'][sheet.cell(row,col1).value]=sheet.cell(row,col3).value.strip() # thuoc tinh tro vao chinh no
                    else:
                        json['Rel'][sheet.cell(row,col1).value]=sheet.cell(row,col3).value.strip() # thuoc tinh tro den doi tuong khac
                
                
                    
        return json
    
        
